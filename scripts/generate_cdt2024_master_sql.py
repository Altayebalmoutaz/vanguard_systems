from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

CATEGORY_SET = {
    "Diagnostic",
    "Preventive",
    "Restorative",
    "Endodontics",
    "Periodontics",
    "Prosthodontics",
    "Oral Surgery",
    "Orthodontics",
    "Adjunctive General Services",
    "Implant Services",
    "Maxillofacial Prosthetics",
}


def esc(value: str) -> str:
    return value.replace("'", "''")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def sql_date(value: date | datetime | None) -> str:
    if value is None:
        return "null"
    if isinstance(value, datetime):
        return f"'{value.date().isoformat()}'"
    return f"'{value.isoformat()}'"


def parse_row(row: tuple[object, ...]) -> tuple[str, str, str | None, str | None, date | datetime | None, str | None, str | None]:
    cells = [clean_text(c) for c in row if c is not None and clean_text(c)]
    code = cells[0] if cells else ""

    # Locate category column dynamically because source sheet has occasional misaligned commas.
    category_idx = None
    for i, c in enumerate(cells[1:], start=1):
        if c in CATEGORY_SET:
            category_idx = i
            break

    if category_idx is None:
        # Fallback: treat second cell as description, unknown category.
        description = cells[1] if len(cells) > 1 else ""
        category = None
        subcategory = None
    else:
        description = ", ".join(cells[1:category_idx]).strip(", ").strip()
        category = cells[category_idx]
        subcategory = cells[category_idx + 1] if len(cells) > category_idx + 1 else None

    effective = None
    for c in row:
        if isinstance(c, (date, datetime)):
            effective = c
            break

    # Pull supplemental narrative fields from later string cells.
    tail = []
    for c in cells:
        if c in (code, description, category or "", subcategory or ""):
            continue
        if c in CATEGORY_SET:
            continue
        if re.fullmatch(r"D\d{4}", c):
            continue
        tail.append(c)
    notes = "; ".join(tail[:2]) if tail else None
    status = None
    return code, description, category, subcategory, effective, status, notes


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    source = Path(r"c:\Users\ZT\Downloads\Telegram Desktop\CDT2024 from PDF.xlsx")
    if not source.exists():
        raise FileNotFoundError(f"Missing source file: {source}")

    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: list[tuple[str, str, str | None, str | None, date | datetime | None, str | None, str | None]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = clean_text(r[0] if r else None).upper()
        if not re.fullmatch(r"D\d{4}", code):
            continue
        parsed = parse_row(r)
        rows.append(parsed)

    out = root / "supabase" / "migrations" / "013_cdt2024_master_load.sql"
    sql: list[str] = []
    sql.append("-- Auto-generated from CDT2024 from PDF.xlsx")
    sql.append("create table if not exists public.cdt_codes (")
    sql.append("  code text primary key,")
    sql.append("  description text not null,")
    sql.append("  category text,")
    sql.append("  subcategory text,")
    sql.append("  effective_date date,")
    sql.append("  status text,")
    sql.append("  notes text,")
    sql.append("  source_file text,")
    sql.append("  updated_at timestamptz not null default now()")
    sql.append(");")
    sql.append("")
    sql.append("truncate table public.cdt_codes;")
    sql.append("")

    batch_size = 400
    for i in range(0, len(rows), batch_size):
        batch = rows[i : i + batch_size]
        sql.append(
            "insert into public.cdt_codes "
            "(code, description, category, subcategory, effective_date, status, notes, source_file) values"
        )
        vals = []
        for code, desc, cat, subcat, eff, status, notes in batch:
            vals.append(
                "("
                f"'{esc(code)}', "
                f"'{esc(desc)}', "
                f"{'null' if cat is None else f"'{esc(cat)}'"}, "
                f"{'null' if subcat is None else f"'{esc(subcat)}'"}, "
                f"{sql_date(eff)}, "
                f"{'null' if status is None else f"'{esc(status)}'"}, "
                f"{'null' if notes is None else f"'{esc(notes)}'"}, "
                "'CDT2024 from PDF.xlsx'"
                ")"
            )
        sql.append(",\n".join(vals))
        sql.append(
            "on conflict (code) do update set "
            "description = excluded.description, "
            "category = excluded.category, "
            "subcategory = excluded.subcategory, "
            "effective_date = excluded.effective_date, "
            "status = excluded.status, "
            "notes = excluded.notes, "
            "source_file = excluded.source_file, "
            "updated_at = now();"
        )
        sql.append("")

    out.write_text("\n".join(sql) + "\n", encoding="utf-8")
    print(f"Wrote {out} with {len(rows)} CDT rows.")


if __name__ == "__main__":
    main()
